#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
from openpyxl import load_workbook
import re
import os
from datetime import datetime

# Constants
EXCEL_FILENAME = "SYSSWRS_Automation.xlsx"
SELECTED_COLUMNS = ['To be tested', 'Accepted', 'Linked by test case', 'Automated']

# Initialize summary lists
processed = []
skipped = []
errors = []

# Load Excel workbook
try:
    wb = load_workbook(EXCEL_FILENAME)
except Exception as e:
    print(f"Error opening Excel file '{EXCEL_FILENAME}': {e}")
    raise SystemExit

# Loop through all CSV files
for csv_filename in os.listdir():
    name_lower = csv_filename.lower()
    if not (name_lower.endswith(".csv") and (name_lower.startswith("sys_") or name_lower.startswith("sw_"))):
        continue

    print(f"Processing file: {csv_filename}")

    # Extract date from filename
    match = re.search(r'(\d{2})_(\d{2})_(\d{4})', csv_filename)
    if not match:
        skipped.append(f"{csv_filename} (date not found)")
        print(f"Skipped {csv_filename} — Date not found in filename.")
        continue
    day, month, year = match.groups()
    date_label = f"{day}-{month}-{year}"

    # Select correct sheet
    ws = wb["SYSSVRS"] if name_lower.startswith("sys_") else wb["SWRS"]

    # Find matching column by date (row 4)
    date_col = None
    for cell in ws[4]:
        # Handle Excel datetime or string
        cell_value = cell.value
        if isinstance(cell_value, datetime):
            # Convert both to date objects for comparison
            excel_date = cell_value.date() if hasattr(cell_value, 'date') else cell_value
            csv_date = datetime.strptime(date_label, "%d-%m-%Y").date()
            print(f"DEBUG: Excel {excel_date}, CSV {csv_date}")  # <--- debug
            if excel_date == csv_date:
                date_col = cell.column_letter
                print(f"Matched Excel date {excel_date} -> Column {date_col}")
                break
        else:
            # Fallback for text headers
            if str(cell_value).strip() == date_label:
                date_col = cell.column_letter
                break

    if date_col is None:
        skipped.append(f"{csv_filename} (no matching column for {date_label})")
        print(f"{csv_filename}: No matching column for {date_label} in Excel — skipped.")
        continue

    # Load CSV and sum columns
    try:
        csv_data = pd.read_csv(csv_filename, on_bad_lines='skip')
        missing = [col for col in SELECTED_COLUMNS if col not in csv_data.columns]
        if missing:
            skipped.append(f"{csv_filename} (missing columns: {missing})")
            print(f"Skipped {csv_filename} — Missing columns: {missing}")
            continue

        column_sum = csv_data[SELECTED_COLUMNS].sum(numeric_only=True)
    except Exception as e:
        errors.append(f"{csv_filename}: {e}")
        print(f"Error reading {csv_filename}: {e}")
        continue

    # Insert data into Excel
    try:
        ws[f"{date_col}6"] = column_sum['To be tested']
        ws[f"{date_col}7"] = column_sum['Accepted']
        ws[f"{date_col}9"] = column_sum['Linked by test case']
        ws[f"{date_col}10"] = column_sum['Automated']
        processed.append(csv_filename)
        print(f"Inserted data from {csv_filename} into column {date_col} ({date_label}) successfully.\n")
    except Exception as e:
        errors.append(f"{csv_filename}: {e}")
        print(f"Error writing {csv_filename} to Excel: {e}")
        continue

# Save Excel workbook
try:
    wb.save(EXCEL_FILENAME)
    wb.close()
except Exception as e:
    print(f"Error saving Excel file: {e}")
    raise SystemExit

# Summary
print("\nSummary Report")
print(f"Processed: {len(processed)} files")
for f in processed: print(f"  - {f}")
print(f"Skipped: {len(skipped)} files")
for f in skipped: print(f"  - {f}")
print(f"Errors: {len(errors)} files")
for f in errors: print(f"  - {f}")

